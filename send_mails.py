# send_mails.py
import csv
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email import encoders
from email.utils import formataddr
from docx import Document
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
# ============ CONFIGURATION ============
GMAIL_USER = 'aniketajawale@gmail.com'
GMAIL_PASS = 'fgopfurasyvtlktx'  # Use Gmail App Password

EMAIL_SUBJECT = "Aniket Jawale - Azure Certified | Application for Data Analytics / Business Analyst / Cloud Data Engineering Roles"

# Define base path
BASE_DIR = r'C:\Users\anike\OneDrive\Desktop\JOB HUNT\Personalized Email Sender'

# ============ EMAIL LOGGING SETUP ============
LOG_FILE = os.path.join(BASE_DIR, "email_log.xlsx")

# Initialize Excel log file if not present
if not os.path.exists(LOG_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sent Emails"
    ws.append(["Date", "Time", "Name", "Email", "Company"])
    wb.save(LOG_FILE)

# ============ LOAD CONTACTS ============
with open(os.path.join(BASE_DIR, 'contacts.csv'), newline='', encoding='ISO-8859-1') as csvfile:
    reader = csv.DictReader(csvfile)
    contacts = [row for row in reader]

# ============ LOAD HTML TEMPLATE ============
with open(os.path.join(BASE_DIR, 'email_template.html'), 'r', encoding='utf-8') as f:
    email_template = f.read()

# ============ SEND EMAILS ============
for contact in contacts:
    if not contact.get('Name') or not contact.get('Email') or not contact.get('Company'):
        print(f"⚠️ Skipping invalid row: {contact}")
        continue

    name = contact['Name'].strip()
    email = contact['Email'].strip()
    company = contact['Company'].strip()

    print(f"📨 Sending email to {name} at {email} ({company})")

    # Personalize HTML body
    html_body = email_template.replace("{{Name}}", name).replace("{{Company}}", company)

    # Create message container
    msg = MIMEMultipart('related')
    msg['From'] = formataddr(("Aniket Jawale", GMAIL_USER))
    msg['To'] = email
    msg['Subject'] = EMAIL_SUBJECT

    # Create HTML body with embedded images
    msg_alternative = MIMEMultipart('alternative')
    msg.attach(msg_alternative)
    msg_alternative.attach(MIMEText(html_body, 'html'))

    # Attach images inline
    for i in range(1, 5):
        image_path = os.path.join(BASE_DIR, f'image{i}.png')
        with open(image_path, 'rb') as img_file:
            img = MIMEImage(img_file.read())
            img.add_header('Content-ID', f'<image{i}>')
            img.add_header('Content-Disposition', 'inline', filename=f'image{i}.png')
            msg.attach(img)

    # Attach common PDFs
    for pdf_name in ['Aniket_Jawale_Resume.pdf', 'Aniket_Projects.pdf']:
        pdf_path = os.path.join(BASE_DIR, pdf_name)
        with open(pdf_path, 'rb') as pdf_file:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(pdf_file.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{pdf_name}"')
            msg.attach(part)

    # Generate personalized DOCX
    docx_template = os.path.join(BASE_DIR, 'base_doc.docx')
    doc = Document(docx_template)
    for p in doc.paragraphs:
        if '{{Name}}' in p.text or '{{Company}}' in p.text:
            for run in p.runs:
                run.text = run.text.replace('{{Name}}', name).replace('{{Company}}', company)

    temp_docx = os.path.join(BASE_DIR, "Aniket_Cover_Letter.docx")
    doc.save(temp_docx)

    with open(temp_docx, 'rb') as file:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(file.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(temp_docx)}"')
        msg.attach(part)

    os.remove(temp_docx)  # Clean up after sending

    # ✅ Send Email after message is fully constructed
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
        server.login(GMAIL_USER, GMAIL_PASS)
        server.sendmail(GMAIL_USER, email, msg.as_string())
        print(f"✅ Email sent to {email}\n")

          # Log after sending
    wb = load_workbook(LOG_FILE)
    ws = wb.active
    now = datetime.now()
    ws.append([now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"), name, email, company])
    wb.save(LOG_FILE)

print("\n🎉 All emails prepared and sent successfully!")
